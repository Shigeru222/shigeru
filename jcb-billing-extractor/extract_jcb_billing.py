"""
JCBメール請求金額抽出スクリプト
mail@qa.jcb.co.jp からのメールを取得し、請求金額をExcelに記録してGoogleドライブに保存する
"""

import os
import re
import base64
import pickle
from datetime import datetime
from email import message_from_bytes
from email.header import decode_header

from google.auth.transport.requests import Request
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from bs4 import BeautifulSoup

SCOPES = [
    "https://www.googleapis.com/auth/gmail.readonly",
    "https://www.googleapis.com/auth/drive.file",
]

TARGET_SENDER = "mail@qa.jcb.co.jp"
EXCEL_FILENAME = "JCB請求履歴.xlsx"
TOKEN_FILE = "token.pickle"
CREDENTIALS_FILE = "credentials.json"

# JCBメール内の請求金額を抽出する正規表現パターン
BILLING_PATTERNS = [
    r"ご請求金額[　\s]*[：:][　\s]*([0-9,，]+)円",
    r"請求金額[　\s]*[：:][　\s]*([0-9,，]+)円",
    r"お支払金額[　\s]*[：:][　\s]*([0-9,，]+)円",
    r"お支払い金額[　\s]*[：:][　\s]*([0-9,，]+)円",
    r"合計金額[　\s]*[：:][　\s]*([0-9,，]+)円",
    r"([0-9,，]+)円.*ご請求",
]

# 請求月を抽出するパターン
BILLING_MONTH_PATTERNS = [
    r"(\d{4})年\s*(\d{1,2})月.*?ご請求",
    r"(\d{4})/(\d{1,2}).*?ご請求",
    r"(\d{4})年(\d{1,2})月分",
]

# 支払日を抽出するパターン
PAYMENT_DATE_PATTERNS = [
    r"(\d{4})年\s*(\d{1,2})月\s*(\d{1,2})日.*?お引き落とし",
    r"お引き落とし日[　\s]*[：:][　\s]*(\d{4})年\s*(\d{1,2})月\s*(\d{1,2})日",
    r"支払日[　\s]*[：:][　\s]*(\d{4})年\s*(\d{1,2})月\s*(\d{1,2})日",
    r"(\d{4})年\s*(\d{1,2})月\s*(\d{1,2})日.*?引落",
]


def authenticate():
    """Gmail / Google Drive の認証を行い、credentialsを返す"""
    creds = None

    if os.path.exists(TOKEN_FILE):
        with open(TOKEN_FILE, "rb") as f:
            creds = pickle.load(f)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(CREDENTIALS_FILE, SCOPES)
            creds = flow.run_local_server(port=0)

        with open(TOKEN_FILE, "wb") as f:
            pickle.dump(creds, f)

    return creds


def decode_subject(subject_raw):
    """メール件名をデコードして文字列として返す"""
    parts = decode_header(subject_raw)
    decoded_parts = []
    for part, charset in parts:
        if isinstance(part, bytes):
            decoded_parts.append(part.decode(charset or "utf-8", errors="replace"))
        else:
            decoded_parts.append(part)
    return "".join(decoded_parts)


def extract_body(msg):
    """メールメッセージからテキスト本文を抽出する"""
    body = ""

    if msg.is_multipart():
        for part in msg.walk():
            content_type = part.get_content_type()
            charset = part.get_content_charset() or "utf-8"

            if content_type == "text/plain":
                payload = part.get_payload(decode=True)
                if payload:
                    body = payload.decode(charset, errors="replace")
                    break
            elif content_type == "text/html" and not body:
                payload = part.get_payload(decode=True)
                if payload:
                    html = payload.decode(charset, errors="replace")
                    soup = BeautifulSoup(html, "html.parser")
                    body = soup.get_text(separator="\n")
    else:
        charset = msg.get_content_charset() or "utf-8"
        payload = msg.get_payload(decode=True)
        if payload:
            raw = payload.decode(charset, errors="replace")
            if msg.get_content_type() == "text/html":
                soup = BeautifulSoup(raw, "html.parser")
                body = soup.get_text(separator="\n")
            else:
                body = raw

    return body


def extract_billing_amount(body):
    """本文から請求金額（数値）を抽出する。見つからない場合は None を返す"""
    for pattern in BILLING_PATTERNS:
        match = re.search(pattern, body)
        if match:
            amount_str = match.group(1).replace(",", "").replace("，", "")
            try:
                return int(amount_str)
            except ValueError:
                continue
    return None


def extract_billing_month(body):
    """本文から請求月を抽出する。見つからない場合は空文字を返す"""
    for pattern in BILLING_MONTH_PATTERNS:
        match = re.search(pattern, body)
        if match:
            year, month = match.group(1), match.group(2)
            return f"{year}年{int(month):02d}月"
    return ""


def extract_payment_date(body):
    """本文から支払日を抽出する。見つからない場合は空文字を返す"""
    for pattern in PAYMENT_DATE_PATTERNS:
        match = re.search(pattern, body)
        if match:
            if len(match.groups()) == 3:
                year, month, day = match.group(1), match.group(2), match.group(3)
                return f"{year}/{int(month):02d}/{int(day):02d}"
    return ""


def fetch_jcb_emails(gmail_service):
    """JCBからのメールを全件取得してリストで返す"""
    query = f"from:{TARGET_SENDER}"
    results = []
    page_token = None

    print(f"  {TARGET_SENDER} のメールを検索中...")

    while True:
        kwargs = {"userId": "me", "q": query, "maxResults": 100}
        if page_token:
            kwargs["pageToken"] = page_token

        response = gmail_service.users().messages().list(**kwargs).execute()
        messages = response.get("messages", [])

        for msg_ref in messages:
            msg_data = (
                gmail_service.users()
                .messages()
                .get(userId="me", id=msg_ref["id"], format="raw")
                .execute()
            )
            raw = base64.urlsafe_b64decode(msg_data["raw"])
            msg = message_from_bytes(raw)

            subject_raw = msg.get("Subject", "（件名なし）")
            subject = decode_subject(subject_raw)
            date_str = msg.get("Date", "")

            # 受信日時をパース
            try:
                from email.utils import parsedate_to_datetime
                received_dt = parsedate_to_datetime(date_str)
                received_str = received_dt.strftime("%Y/%m/%d %H:%M")
            except Exception:
                received_str = date_str

            body = extract_body(msg)
            billing_amount = extract_billing_amount(body)
            billing_month = extract_billing_month(body)
            payment_date = extract_payment_date(body)

            results.append(
                {
                    "received": received_str,
                    "subject": subject,
                    "sender": TARGET_SENDER,
                    "billing_month": billing_month,
                    "payment_date": payment_date,
                    "billing_amount": billing_amount,
                }
            )
            print(
                f"  取得: {received_str} | {subject[:30]} | "
                f"請求金額: {billing_amount:,}円" if billing_amount else
                f"  取得: {received_str} | {subject[:30]} | 請求金額: 抽出不可"
            )

        page_token = response.get("nextPageToken")
        if not page_token:
            break

    # 受信日時で昇順ソート
    results.sort(key=lambda x: x["received"])
    return results


def create_excel(records):
    """レコードリストからExcelファイルを作成する"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "JCB請求履歴"

    headers = ["受信日時", "件名", "送信元", "請求月", "支払日", "請求金額（円）"]
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row, record in enumerate(records, start=2):
        ws.cell(row=row, column=1, value=record["received"])
        ws.cell(row=row, column=2, value=record["subject"])
        ws.cell(row=row, column=3, value=record["sender"])
        ws.cell(row=row, column=4, value=record["billing_month"])
        ws.cell(row=row, column=5, value=record["payment_date"])

        amount_cell = ws.cell(row=row, column=6, value=record["billing_amount"])
        if record["billing_amount"] is not None:
            amount_cell.number_format = "#,##0"

        if row % 2 == 0:
            row_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = row_fill

    # 列幅の自動調整
    column_widths = [20, 50, 25, 12, 14, 18]
    for col, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    ws.freeze_panes = "A2"

    wb.save(EXCEL_FILENAME)
    print(f"\nExcelファイルを作成しました: {EXCEL_FILENAME}")
    return EXCEL_FILENAME


def upload_to_drive(drive_service, filepath):
    """ExcelファイルをGoogleドライブにアップロードする（同名ファイルは上書き）"""
    filename = os.path.basename(filepath)
    mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    # 既存ファイルを検索
    results = (
        drive_service.files()
        .list(
            q=f"name='{filename}' and trashed=false",
            fields="files(id, name)",
        )
        .execute()
    )
    existing = results.get("files", [])

    media = MediaFileUpload(filepath, mimetype=mime_type, resumable=True)

    if existing:
        file_id = existing[0]["id"]
        drive_service.files().update(
            fileId=file_id,
            media_body=media,
        ).execute()
        print(f"Googleドライブの既存ファイルを更新しました (ID: {file_id})")
        return file_id
    else:
        file_metadata = {"name": filename}
        created = (
            drive_service.files()
            .create(body=file_metadata, media_body=media, fields="id")
            .execute()
        )
        file_id = created.get("id")
        print(f"Googleドライブに新規アップロードしました (ID: {file_id})")
        return file_id


def main():
    print("=== JCB請求メール抽出ツール ===\n")
    print("認証中...")
    creds = authenticate()

    gmail_service = build("gmail", "v1", credentials=creds)
    drive_service = build("drive", "v3", credentials=creds)

    print("\nメール取得中...")
    records = fetch_jcb_emails(gmail_service)

    if not records:
        print("対象メールが見つかりませんでした。")
        return

    print(f"\n合計 {len(records)} 件のメールを取得しました。")

    excel_path = create_excel(records)

    print("\nGoogleドライブにアップロード中...")
    file_id = upload_to_drive(drive_service, excel_path)

    print(f"\n完了！")
    print(f"Googleドライブ: https://drive.google.com/file/d/{file_id}/view")


if __name__ == "__main__":
    main()
